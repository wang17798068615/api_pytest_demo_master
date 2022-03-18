import ast
import json

import openpyxl

from api_pytest_demo_master.data.base_data import DATACESHIYONGLI_PATH
from api_pytest_demo_master.public.request_util.Variable import Variable


class ReadXlsx:
    def __init__(self, filename):
        # xlsx文件绝对路径名称
        self.filename = filename
        # 获取 工作簿对象
        self.workbook = openpyxl.load_workbook(fr"{filename}")

    # 获取所有工作表的表名
    def sheetnames(self):
        try:
            sheetnames = self.workbook.sheetnames
            return sheetnames
        except Exception as e:
            return "输入格式有误:", e

    # 根据表名/索引值获取工作表对象
    def worksheet(self, name=None, num=None):
        """

        :param name: 表名，str类型
        :param num: 索引值，int类型
        :return:
        """
        try:
            if name and isinstance(name, str):
                worksheet = self.workbook[name]
                return worksheet

            elif num == 0:
                shenames = self.sheetnames()
                worksheet = self.workbook[shenames[0]]
                return worksheet

            elif num and isinstance(num, int):
                shenames = self.sheetnames()
                worksheet = self.workbook[shenames[num]]
                return worksheet

            else:
                return "输入格式有误！"
        except Exception as e:
            return "输入格式有误:", e

    # 获取工作表的表名title
    def title(self, name=None, num=None):
        try:
            worksheet = self.worksheet(name=name, num=num)
            res_name = worksheet.title
            return res_name
        except Exception as e:
            return "输入格式有误:", e

    # 获取工作表的“行数”max_row
    def max_row(self, name=None, num=None):
        try:
            worksheet = self.worksheet(name=name, num=num)
            rows = worksheet.max_row
            return rows
        except Exception as e:
            return "输入格式有误:", e

    # 获取工作表的“列数”max_column
    def max_column(self, name=None, num=None):
        try:
            worksheet = self.worksheet(name=name, num=num)
            columns = worksheet.max_column
            return columns
        except Exception as e:
            return "输入格式有误:", e

    # 按行方式获取表中的所有数据(格式为：[['北京', ' 石家庄'], [1, 2]]) --> ["参数":"参数"] ，["参数值":"参数值"]
    def rows(self, name=None, num=None):
        try:
            res_list = []
            worksheet = self.worksheet(name=name, num=num)
            for row in worksheet.rows:
                list = []
                for cell in row:
                    list.append(cell.value)
                res_list.append(list)
            return res_list
        except Exception as e:
            return "输入格式有误:", e

    # 按列方式获取表中的所有数据(格式为:[['北京', 1], [' 石家庄', 2]]) --> ["参数":"参数值"]，注：只适用于俩行数据，不推荐使用
    def columns(self, name=None, num=None):
        try:
            res_list = []
            worksheet = self.worksheet(name=name, num=num)
            for col in worksheet.columns:
                list = []
                for cell in col:
                    list.append(cell.value)
                res_list.append(list)
            return res_list
        except Exception as e:
            return "输入格式有误:", e

    # 获取表中几行几列的数据，如何没有数据就会返回None【注意：在Excel表，行和列都是从1开始计数的!!】
    def cells(self, row, column, name=None, num=None):
        """

        :param row: 行，int类型
        :param column: 列，int类型
        :return:
        """
        try:
            res_list = []
            worksheet = self.worksheet(name=name, num=num)
            for i in range(1, row + 1):
                list = []
                for j in range(1, column + 1):
                    value = worksheet.cell(row=i, column=j).value
                    list.append(value)
                res_list.append(list)
            return res_list
        except Exception as e:
            return "输入格式有误:", e

    # 获取表中某一单元格的数据
    # 精确读取表格中的某一单元格
    def value(self, pos="A1", name=None, num=None):
        """

        :param pos: xlsx坐标值，例：A1
        :return:
        """
        try:
            worksheet = self.worksheet(name=name, num=num)
            result = worksheet[pos].value
            return result
        except Exception as e:
            return "输入格式有误:", e

    # 获取表中某一单元格的数据，并重新覆盖写入新值
    def writeValue(self, pos="A1", values=None, name=None, num=None):
        """

        :param pos: xlsx坐标值，例：A1
        :param values: 写入的值，str类型
        :return:
        """
        try:
            worksheet = self.worksheet(name=name, num=num)
            worksheet[pos].value = values
            # 一旦做了修改，就要保存，保存的时候，要保证没有其它程序在使用当前文件。否则会报Permission Error
            self.workbook.save(self.filename)
            return f"写入成功,写入的值为:{values}"
        except Exception as e:
            return "输入格式有误:", e

    # 将行方式获取xlsx表数据整合成列表嵌套字典类型
    def getListDictRowsXlsx(self, name=None, num=None):
        reader = self.rows(name=name, num=num)
        # [{'csid':1,'page':1},{}]
        dictList = []
        for i in range(1, len(reader)):
            dict = {}  # 一条数据的字典
            data = reader[i]  # 一条csv数据
            keys = reader[0]  # key数据
            for j in range(len(keys)):
                dict[keys[j]] = data[j]  # 组装一条字典
            dictList.append(dict)  # 添加到列表

        return dictList

    # 将获取表中几行几列的xlsx表数据整合成列表嵌套字典类型
    def getListDictCellsXlsx(self, row, column, name=None, num=None):
        reader = self.cells(row, column, name=name, num=num)
        # [{'csid':1,'page':1},{}]
        dictList = []
        for i in range(1, len(reader)):
            dict = {}  # 一条数据的字典
            data = reader[i]  # 一条csv数据
            keys = reader[0]  # key数据
            for j in range(len(keys)):
                dict[keys[j]] = data[j]  # 组装一条字典

            dictList.append(dict)  # 添加到列表

        return dictList


if __name__ == '__main__':
    # r = ReadXlsx("D:\ww.xlsx")
    # r = ReadXlsx(r"D:\测试用例.xlsx")
    r = ReadXlsx(r"D:\pycharm_workspace\proTest\api_pytest_demo_master\data\测试用例.xlsx")
    print(r.getListDictRowsXlsx(name="全局变量"))
    print(r.getListDictRowsXlsx(num=5))

'''
[{'key': 'user', 'value': 15500000000}, {'key': 'passwd', 'value': 12345678}, {'key': 'admin', 'value': 15500000011}, {'key': 'admin_pwd', 'value': '{"a":"b"}'}, {'key': 'mock_url', 'value': 'https://www.fastmock.site/mock/206dd047d2278edb07eeaf592905e74d/nmb'}]
[{
'id': 1,
 'title': '登录成功-普通用户', 
'pre_sql': '{"use_database":"use futureloan","sql":"update member set  leave_amount=\'8000\' WHERE mobile_phone=\'13927427491\';"}',
 'method': 'post', 
 'url': '/member/login', 
 'headers': '{"X-Lemonban-Media-Type": "lemonban.v2"}',
  'req_data': '{"mobile_phone": "#user#","pwd": "#passwd#"}',
   'assert_list': '[{"expr":"$.code","expected":0,"type":"eq"},\n{"expr":"$.msg","expected":"OK","type":"eq"},\n{"expr":"$..mobile_phone","expected":"#user#","type":"eq"}]', 'extract': '{"msg":"$.msg"}', 
   'assert_db': '[{"use_database":"use futureloan","sql":"select id from member where mobile_phone=\'#user#\'","expected":1,"db_type":"count"}\n]'
   }]

'''
